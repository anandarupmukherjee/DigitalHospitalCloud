import json
from datetime import timedelta
from html.parser import HTMLParser
from io import BytesIO
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.test import SimpleTestCase, TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from services.data_collection_tray.listener import TrayMQTTListener
from services.data_storage.repository import record_tray_heartbeat

from .models import TrayEvent, TrayStatus


class _CanvasDataParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.values = {}

    def handle_starttag(self, tag, attrs):
        if tag != "canvas":
            return
        attr_map = dict(attrs)
        element_id = attr_map.get("id")
        if not element_id or element_id in self.values:
            return
        data_attr = attr_map.get("data-chart")
        if data_attr is not None:
            self.values[element_id] = data_attr


class TrayHistoryViewTests(TestCase):

    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(
            username="manager",
            password="password",
            is_superuser=True,
            is_staff=True,
        )
        self.tray = TrayStatus.objects.create(
            tray_id="TRAY-123",
            topic="hospital/tray/TRAY-123",
            is_active=False,
        )

    def test_carry_over_session_counts_toward_chart_data(self):
        now = timezone.now()
        on_time = now - timedelta(days=1, hours=2)
        off_time = now - timedelta(hours=5)
        TrayEvent.objects.create(
            tray=self.tray,
            status=TrayEvent.STATUS_ON,
            timestamp=on_time,
        )
        TrayEvent.objects.create(
            tray=self.tray,
            status=TrayEvent.STATUS_OFF,
            timestamp=off_time,
        )

        self.client.force_login(self.user)
        with patch("tracker.views.timezone.now", return_value=now):
            response = self.client.get(
                reverse("tray-history"), {"tray": self.tray.id, "range": "day"}
            )

        self.assertEqual(response.status_code, 200)
        activation_periods = response.context["activation_periods"]
        self.assertEqual(len(activation_periods), 1)
        period = activation_periods[0]
        window_start = now - timedelta(days=1)
        self.assertEqual(period["start"], window_start)
        expected_minutes = (off_time - window_start).total_seconds() / 60
        self.assertAlmostEqual(period["duration_minutes"], expected_minutes, places=5)

        chart_data = json.loads(response.context["chart_data_json"])
        self.assertEqual(chart_data["data"], [round(expected_minutes, 2)])

        parser = _CanvasDataParser()
        parser.feed(response.content.decode())
        history_attr = parser.values.get("trayHistoryChart")
        self.assertIsNotNone(history_attr)
        self.assertNotIn("\\u0022", history_attr)
        self.assertEqual(json.loads(history_attr), chart_data)

        summary_context = json.loads(response.context["summary_chart_data_json"])
        summary_attr = parser.values.get("traySummaryChart")
        self.assertIsNotNone(summary_attr)
        self.assertNotIn("\\u0022", summary_attr)
        self.assertEqual(json.loads(summary_attr), summary_context)

    def test_delete_tray_removes_records(self):
        now = timezone.now()
        TrayEvent.objects.create(
            tray=self.tray,
            status=TrayEvent.STATUS_ON,
            timestamp=now - timedelta(hours=2),
        )
        TrayEvent.objects.create(
            tray=self.tray,
            status=TrayEvent.STATUS_OFF,
            timestamp=now - timedelta(hours=1),
        )

        self.client.force_login(self.user)
        response = self.client.post(reverse("tray-history"), {"tray": self.tray.id})

        self.assertRedirects(response, reverse("tray-history"))
        self.assertFalse(TrayStatus.objects.filter(id=self.tray.id).exists())
        self.assertEqual(TrayEvent.objects.filter(tray=self.tray).count(), 0)

    def test_tray_history_includes_heartbeat_context(self):
        now = timezone.now()
        earlier = now - timedelta(seconds=8)
        record_tray_heartbeat(
            self.tray.tray_id,
            topic="MET/hospital/status/TRAY-123",
            event_time=earlier,
            payload={"tray_id": self.tray.tray_id},
        )
        record_tray_heartbeat(
            self.tray.tray_id,
            topic="MET/hospital/status/TRAY-123",
            event_time=now,
            payload={"tray_id": self.tray.tray_id},
        )

        self.client.force_login(self.user)
        with patch("tracker.views.timezone.now", return_value=now):
            response = self.client.get(
                reverse("tray-history"), {"tray": self.tray.id, "range": "day"}
            )

        self.assertEqual(response.status_code, 200)
        heartbeat_status = response.context["heartbeat_status"]
        self.assertTrue(heartbeat_status["is_alive"])
        heartbeat_events = response.context["heartbeat_events"]
        self.assertGreaterEqual(len(heartbeat_events), 2)
        self.assertIn("down", {event.status for event in heartbeat_events})


class DashboardViewTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(
            username="manager_dash",
            password="password",
            is_superuser=True,
            is_staff=True,
        )
        self.tray_a = TrayStatus.objects.create(tray_id="TRAY-A", is_active=False)
        self.tray_b = TrayStatus.objects.create(tray_id="TRAY-B", is_active=False)

    def _create_session(self, tray, start, end):
        TrayEvent.objects.create(tray=tray, status=TrayEvent.STATUS_ON, timestamp=start)
        TrayEvent.objects.create(tray=tray, status=TrayEvent.STATUS_OFF, timestamp=end)

    def test_dashboard_chart_payload_contains_tray_stats(self):
        now = timezone.now()
        self._create_session(
            self.tray_a,
            now - timedelta(days=2, hours=2),
            now - timedelta(days=2, hours=1, minutes=30),
        )
        self._create_session(
            self.tray_a,
            now - timedelta(hours=5),
            now - timedelta(hours=4, minutes=15),
        )
        self._create_session(
            self.tray_b,
            now - timedelta(hours=10),
            now - timedelta(hours=9, minutes=30),
        )

        self.client.force_login(self.user)
        with patch("tracker.views.timezone.now", return_value=now):
            response = self.client.get(reverse("dashboard"))

        self.assertEqual(response.status_code, 200)
        chart_json = response.context["dashboard_summary_chart_json"]
        self.assertTrue(chart_json)
        chart_data = json.loads(chart_json)
        self.assertEqual(chart_data["labels"], ["TRAY-A", "TRAY-B"])
        self.assertEqual(len(chart_data["avg_minutes"]), 2)

        parser = _CanvasDataParser()
        parser.feed(response.content.decode())
        canvas_value = parser.values.get("dashboardSummaryChart")
        self.assertIsNotNone(canvas_value)
        self.assertNotIn("\\u0022", canvas_value)
        self.assertEqual(json.loads(canvas_value), chart_data)


class TrayStatusDataViewTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(
            username="heartbeat_user",
            password="password",
            is_superuser=True,
            is_staff=True,
        )
        self.tray = TrayStatus.objects.create(
            tray_id="TRAY-API",
            topic="hospital/status/TRAY-API",
            latitude=12.34,
            longitude=56.78,
            is_active=False,
        )
        record_tray_heartbeat(
            self.tray.tray_id,
            topic=self.tray.topic,
            event_time=timezone.now(),
            payload={"tray_id": self.tray.tray_id},
        )

    def test_api_returns_heartbeat_summary(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse("tray-status-api"))
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        heartbeat = payload["heartbeat"]
        summary = heartbeat["summary"]
        self.assertEqual(summary["tray_id"], self.tray.tray_id)
        self.assertTrue(heartbeat["records"])

    def test_api_excludes_unknown_tray_heartbeats(self):
        record_tray_heartbeat(
            "UNKNOWN",
            topic="MET/hospital/status/UNKNOWN",
            event_time=timezone.now(),
            payload={"tray_id": "UNKNOWN"},
        )
        self.client.force_login(self.user)
        response = self.client.get(reverse("tray-status-api"))
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        ids = {record["tray_id"] for record in payload["heartbeat"]["records"]}
        self.assertIn(self.tray.tray_id, ids)
        self.assertNotIn("UNKNOWN", ids)


class TrayReportDownloadViewTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user(
            username="report_user",
            password="password",
            is_superuser=True,
            is_staff=True,
        )
        self.tray = TrayStatus.objects.create(
            tray_id="TRAY-REPORT",
            topic="hospital/tray/TRAY-REPORT",
            is_active=False,
        )

    def test_report_contains_collection_and_heartbeat_sheets(self):
        now = timezone.now()
        TrayEvent.objects.create(
            tray=self.tray,
            status=TrayEvent.STATUS_ON,
            timestamp=now - timedelta(hours=2),
        )
        TrayEvent.objects.create(
            tray=self.tray,
            status=TrayEvent.STATUS_OFF,
            timestamp=now - timedelta(hours=1, minutes=15),
        )
        record_tray_heartbeat(
            self.tray.tray_id,
            topic="MET/hospital/status/TRAY-REPORT",
            event_time=now - timedelta(seconds=7),
            payload={"tray_id": self.tray.tray_id},
        )
        record_tray_heartbeat(
            self.tray.tray_id,
            topic="MET/hospital/status/TRAY-REPORT",
            event_time=now,
            payload={"tray_id": self.tray.tray_id},
        )

        self.client.force_login(self.user)
        with patch("tracker.views.timezone.now", return_value=now):
            response = self.client.get(
                reverse("tray-report-download", args=[self.tray.id])
            )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response["Content-Type"],
        )
        workbook = load_workbook(filename=BytesIO(response.content))
        self.assertIn("collection", workbook.sheetnames)
        self.assertIn("alive", workbook.sheetnames)
        self.assertGreater(workbook["collection"].max_row, 1)
        self.assertGreater(workbook["alive"].max_row, 1)


class TrayMQTTListenerTests(SimpleTestCase):
    def test_status_topic_suffix_used_as_tray_id(self):
        listener = TrayMQTTListener(status_topic="MET/hospital/status/#")
        topic = "MET/hospital/status/TRAY-ABC/heartbeat"

        with patch("services.data_collection_tray.listener.record_tray_heartbeat") as mock_record:
            listener._handle_heartbeat(topic, payload={})

        mock_record.assert_called_once()
        tray_id_arg = mock_record.call_args[0][0]
        self.assertEqual(tray_id_arg, "TRAY-ABC")
